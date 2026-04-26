from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.utils import timezone
from .models import Santri, Presensi, SuratIzin, RegistrationCode
from .serializers import SantriSerializer, SuratIzinSerializer, UserSerializer, RegisterSantriAccountSerializer
from .face_utils import decode_base64_image, recognize_from_image_pil, recognize_many_from_image_pil, encode_face_from_image, invalidate_face_model_cache
from django.contrib.auth.models import User
from rest_framework.authtoken.models import Token
from rest_framework import generics, status
from django.core.cache import cache
import datetime
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from PIL import Image
from django.db import IntegrityError
from openpyxl.styles import Alignment, Font, PatternFill
from .utils import get_rekap_data
import os
from django.conf import settings


def _replace_presensi_record(santri, tanggal, sesi, kelas, status_presensi, created_by):
    qs = Presensi.objects.filter(
        santri=santri,
        tanggal=tanggal,
        sesi=sesi,
        kelas=kelas,
    ).order_by('-id')

    existing = qs.first()
    if existing:
        existing.status = status_presensi
        existing.kelas = kelas
        existing.created_by = created_by
        existing.waktu_scan = timezone.now()
        existing.save(update_fields=['status', 'kelas', 'created_by', 'waktu_scan'])
        qs.exclude(id=existing.id).delete()
        return existing

    return Presensi.objects.create(
        santri=santri,
        tanggal=tanggal,
        sesi=sesi,
        kelas=kelas,
        status=status_presensi,
        created_by=created_by,
        waktu_scan=timezone.now(),
    )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_logout(request):
    try:
        if hasattr(request.user, "auth_token"):
            request.user.auth_token.delete()
    except Exception:
        pass
    return Response({"ok": True, "message": "Logout berhasil"})

PRESENSI_KEY = "presensi_start_time"
TELAT_KEY = "telat_start_time"

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_permohonan_izin(request):
    santri_id = request.data.get('santri')
    tanggal = request.data.get('tanggal')
    sesi = request.data.get('sesi')
    alasan = request.data.get('alasan')
    kelas = request.data.get('kelas')
    
    if not (santri_id and tanggal and sesi and alasan):
        return Response({'ok': False, 'message': 'Lengkapi data'}, status=400)
    
    try:
        s = Santri.objects.get(santri_id=santri_id)
    except Santri.DoesNotExist:
        try:
            s = Santri.objects.get(pk=santri_id)
        except Santri.DoesNotExist:
            return Response({'ok': False, 'message': 'Santri tidak ditemukan'}, status=404)
    
    try:
        si = SuratIzin(santri=s, kelas=kelas, tanggal=tanggal, sesi=sesi, alasan=alasan, status='Menunggu')
        si.save()
    except IntegrityError:
        return Response({'ok': False, 'message': 'Sudah ada surat izin untuk santri ini pada tanggal & sesi tersebut'}, status=400)
    
    return Response({'ok': True, 'surat': SuratIzinSerializer(si).data})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_izin_santri(request):
    izin = SuratIzin.objects.filter(santri__user=request.user).select_related('santri').order_by('-tanggal')
    return Response({'ok': True, 'data': SuratIzinSerializer(izin, many=True).data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_permohonan_izin(request):
    user = request.user
    
    if user.is_staff:
        izin = SuratIzin.objects.filter(status="Menunggu").select_related('santri').order_by('-tanggal')
    else:
        santri = Santri.objects.get(user=user)
        izin = SuratIzin.objects.filter(santri=santri).select_related('santri').order_by('-tanggal')

    data = [{
        "id": i.id,
        "santri_id": i.santri.santri_id,
        "nama": i.santri.nama,
        "kelas": i.kelas,
        "tanggal": i.tanggal,
        "sesi": i.sesi,
        "alasan": i.alasan,
        "status": i.status
    } for i in izin]
    return Response({"ok": True, "data": data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def validasi_izin(request, izin_id):
    try:
        izin = SuratIzin.objects.get(id=izin_id)
    except SuratIzin.DoesNotExist:
        return Response({'ok': False, 'message': 'Izin tidak ditemukan'}, status=404)

    action = request.data.get('status')
    if action not in ['Disetujui', 'Ditolak']:
        return Response({'ok': False, 'message': 'Status tidak valid'}, status=400)

    izin.status = action
    izin.note = request.data.get('note', '')
    izin.save()
    return Response({'ok': True, 'message': f'Izin {action} berhasil'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_list_santri(request):
    santris = Santri.objects.all().order_by('santri_id')
    return Response({'ok': True, 'data': SantriSerializer(santris, many=True).data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_search_santri_for_deactivation(request):
    if not request.user.is_staff:
        return Response({'ok': False, 'message': 'Akses ditolak'}, status=403)

    nama = (request.query_params.get('nama') or '').strip()
    asal_daerah = (request.query_params.get('asal_daerah') or '').strip()

    if not nama or not asal_daerah:
        return Response(
            {'ok': False, 'message': 'Nama lengkap dan asal daerah wajib diisi'},
            status=400
        )

    santris = Santri.objects.filter(
        nama__icontains=nama,
        asal_daerah__icontains=asal_daerah
    ).order_by('nama', 'santri_id')

    if not santris.exists():
        return Response({'ok': False, 'message': 'Santri tidak ditemukan', 'data': []}, status=404)

    return Response({'ok': True, 'data': SantriSerializer(santris, many=True).data})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_deactivate_santri(request, santri_id):
    if not request.user.is_staff:
        return Response({'ok': False, 'message': 'Akses ditolak'}, status=403)

    try:
        santri = Santri.objects.select_related('user').get(id=santri_id)
    except Santri.DoesNotExist:
        return Response({'ok': False, 'message': 'Santri tidak ditemukan'}, status=404)

    santri_name = santri.nama
    user = santri.user
    if user:
        user.delete()
    else:
        santri.delete()

    RegistrationCode.objects.filter(santri_name__iexact=santri_name, used=False).delete()
    invalidate_face_model_cache()

    return Response({'ok': True, 'message': 'Santri berhasil dinonaktifkan dan dihapus dari database'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_get_user(request):
    user = request.user
    role = "pengurus" if user.is_staff else "santri"
    santri_profile = getattr(user, "santri_profile", None)
    full_name = santri_profile.nama if santri_profile else getattr(user, "full_name", "")

    return Response({
        "id": user.id,
        "username": user.username,
        "full_name": full_name,
        "is_staff": user.is_staff,
        "is_superuser": user.is_superuser,
        "role": role
    })

class RegisterPengurusView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

class RegisterSantriView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSantriAccountSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        registration_code = request.data.get('registration_code', '').strip()
        if not registration_code:
            return Response({
                "ok": False,
                "message": "Kode registrasi wajib diisi. Hubungi pengurus untuk mendapatkan kode."
            }, status=400)
        
        try:
            reg_code = RegistrationCode.objects.get(code=registration_code)
        except RegistrationCode.DoesNotExist:
            return Response({
                "ok": False,
                "message": "Kode registrasi tidak valid"
            }, status=400)
        
        if reg_code.used:
            return Response({
                "ok": False,
                "message": "Kode registrasi sudah pernah digunakan"
            }, status=400)
        
        if not reg_code.is_valid():
            return Response({
                "ok": False,
                "message": "Kode registrasi sudah kadaluarsa"
            }, status=400)

        nama_input = request.data.get('nama', '').strip()
        if nama_input.lower() != reg_code.santri_name.lower():
            return Response({
                "ok": False,
                "message": f"Nama tidak sesuai dengan kode registrasi. Kode ini untuk: {reg_code.santri_name}"
            }, status=400)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        santri = serializer.save()
        
        reg_code.used = True
        reg_code.used_by = santri.user
        reg_code.save()

        return Response({
            "ok": True,
            "santri_id": santri.id,
            "user_id": santri.user.id,
            "username": santri.user.username,
            "nama": santri.nama,
            "sektor": santri.sektor,
            "role": "santri",
            "message": "Registrasi berhasil!"
        })

class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        from django.contrib.auth import authenticate
        username = request.data.get("username")
        password = request.data.get("password")

        if not username or not password:
            return Response({"error": "Username dan password harus diisi"}, status=400)
        
        user = authenticate(username=username, password=password)
        if not user:
            return Response({"error": "Username atau password salah"}, status=401)

        token, _ = Token.objects.get_or_create(user=user)
        role = "pengurus" if user.is_staff else "santri"
        santri_profile = getattr(user, "santri_profile", None)

        return Response({
            "token": token.key,
            "role": role,
            "user": {
                "id": user.id,
                "username": user.username,
                "full_name": santri_profile.nama if santri_profile else getattr(user, "full_name", ""),
                "santri_id": santri_profile.id if santri_profile else None,
                "sektor": santri_profile.sektor if santri_profile else None,
                "angkatan": santri_profile.angkatan if santri_profile else None
            }
        })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_santri_registrasi_wajah(request):
    try:
        santri_id = request.data.get("santri_id")
        image_data = request.data.get("image")
        images_data = request.data.get("images")

        if not santri_id:
            santri_profile = getattr(request.user, "santri_profile", None)
            if santri_profile:
                santri_id = santri_profile.id

        has_images_list = isinstance(images_data, list) and len(images_data) > 0
        if not santri_id:
            return Response(
                {"error": "santri_id tidak ditemukan. Gunakan akun santri atau kirim santri_id."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not image_data and not has_images_list:
            return Response(
                {"error": "image atau images wajib diisi"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            santri = Santri.objects.get(id=santri_id)
        except Santri.DoesNotExist:
            try:
                santri = Santri.objects.get(user_id=santri_id)
            except Santri.DoesNotExist:
                return Response({"error": f"Santri dengan id {santri_id} tidak ditemukan"}, status=404)

        if has_images_list:
            sample_images = images_data[:5]
        else:
            sample_images = [image_data]

        encodings = []
        last_location = None
        failed_samples = 0

        for sample in sample_images:
            if not sample:
                failed_samples += 1
                continue

            pil_img = decode_base64_image(sample)
            hog_encoding, face_location = encode_face_from_image(pil_img)

            if face_location == "detector_unavailable":
                return Response({"error": "Detector wajah tidak tersedia di server"}, status=500)

            if hog_encoding is None:
                failed_samples += 1
                continue

            encodings.append(hog_encoding)
            last_location = face_location

        if len(encodings) == 0:
            return Response({"error": "Wajah tidak ditemukan pada semua pose"}, status=400)

        # Replace old face encoding with latest registration samples.
        santri.face_encoding = encodings if len(encodings) > 1 else encodings[0]
        santri.save()
        invalidate_face_model_cache()

        top, right, bottom, left = last_location if last_location else (0, 0, 0, 0)

        return Response({
            "ok": True,
            "message": f"Wajah {santri.nama} berhasil diregistrasi ({len(encodings)}/{len(sample_images)} pose).",
            "nama": santri.nama,
            "captured": len(encodings),
            "total": len(sample_images),
            "failed": failed_samples,
            "location": {"top": top, "right": right, "bottom": bottom, "left": left}
        })

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_santri_upload_foto(request):
    try:
        santri_id = request.data.get("santri_id")
        foto_file = request.FILES.get("foto")

        if not santri_id or not foto_file:
            return Response({"error": "santri_id dan foto wajib diisi"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            santri = Santri.objects.get(id=santri_id)
        except Santri.DoesNotExist:
            return Response({"error": "Santri tidak ditemukan"}, status=status.HTTP_404_NOT_FOUND)

        santri.foto = foto_file
        santri.save()

        img_path = santri.foto.path
        pil_img = Image.open(img_path).convert("RGB")
        
        hog_encoding, face_location = encode_face_from_image(pil_img)

        if face_location == "detector_unavailable":
            return Response({"error": "Detector wajah tidak tersedia di server"}, status=500)

        if hog_encoding is None:
            return Response({"error": "Wajah tidak terdeteksi, coba gunakan foto yang lebih jelas"}, status=status.HTTP_400_BAD_REQUEST)

        # Replace old face encoding with latest uploaded photo encoding.
        santri.face_encoding = hog_encoding
        santri.save()
        invalidate_face_model_cache()

        return Response({"success": True, "message": "Foto berhasil diupload & encoding disimpan"}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": f"Error proses wajah: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_start_presensi(request):
    tanggal = request.data.get("tanggal")
    sesi = request.data.get("sesi")
    if not tanggal or not sesi:
        return Response({"ok": False, "message": "Lengkapi tanggal & sesi"})
    presensi_info = {
        "tanggal": tanggal,
        "sesi": sesi,
        "time": timezone.now().isoformat(),
        "telat_start": None
    }
    cache.set(PRESENSI_KEY, presensi_info, 3600)
    return Response({"ok": True, "message": "Presensi dimulai"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_start_telat(request):
    presensi_info = cache.get(PRESENSI_KEY)
    if not presensi_info:
        return Response({"ok": False, "message": "Presensi belum dimulai"}, status=400)

    presensi_info["telat_start"] = timezone.now().isoformat()
    cache.set(PRESENSI_KEY, presensi_info, 3600)
    return Response({"ok": True, "message": "Penghitungan keterlambatan dimulai"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_end_presensi(request):
    cache.delete(PRESENSI_KEY)
    cache.delete(TELAT_KEY)
    return Response({"ok": True, "message": "Presensi selesai"})



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_recognize_and_attend(request):
    try:
        data_url = request.data.get('image')
        if not data_url:
            return Response({"ok": False, "message": "Image data tidak ditemukan"}, status=400)
            
        presensi_info = cache.get(PRESENSI_KEY)
        kelas = request.data.get("kelas") or "Kamu kelas apa?"
        claimed_santri_id_raw = request.data.get("claimed_santri_id")
        claimed_full_name_raw = request.data.get("claimed_full_name")
        claimed_santri_id = None

        if claimed_santri_id_raw not in (None, "", "null"):
            try:
                claimed_santri_id = int(claimed_santri_id_raw)
            except (TypeError, ValueError):
                return Response({"ok": False, "message": "claimed_santri_id harus berupa angka"}, status=400)

        if claimed_santri_id is None and claimed_full_name_raw not in (None, ""):
            claimed_full_name = str(claimed_full_name_raw).strip()
            if not claimed_full_name:
                return Response({"ok": False, "message": "Nama lengkap tidak boleh kosong", "info": "invalid_claimed_name"}, status=400)

            matching_santri = Santri.objects.filter(nama__iexact=claimed_full_name).order_by('id')
            match_count = matching_santri.count()

            if match_count == 0:
                return Response({
                    "ok": False,
                    "message": "Nama lengkap tidak ditemukan. Cek lagi penulisan nama.",
                    "info": "invalid_claimed_name",
                    "can_claim_identity": True,
                }, status=400)

            if match_count > 1:
                return Response({
                    "ok": False,
                    "message": "Nama lengkap tidak unik. Gunakan nama yang lebih spesifik.",
                    "info": "duplicate_claimed_name",
                    "can_claim_identity": True,
                }, status=400)

            claimed_santri_id = matching_santri.first().id

        if not presensi_info:
            return Response({"ok": False, "message": "Presensi belum dimulai"}, status=400)

        tanggal = presensi_info['tanggal']
        sesi = presensi_info['sesi']
        telat_start = presensi_info.get('telat_start')

        pil_img = decode_base64_image(data_url)
        recognized, info, rejected, all_locations = recognize_many_from_image_pil(
            pil_img,
            min_prob=0.56,
            max_faces=5,
            claimed_pk=claimed_santri_id,
        )
        if not recognized:
            error_map = {
                "no_face": "Wajah tidak terdeteksi, pastikan wajah terlihat jelas",
                "no_dataset": "Belum ada dataset wajah yang terdaftar",
                "not_enough_classes": "Minimal dua santri perlu diregistrasi agar model SVM bisa dilatih",
                "low_confidence": "Wajah tidak cocok (confidence terlalu rendah)",
                "ambiguous_face": "Wajah mirip dengan santri lain. Lakukan scan satu per satu atau isi di form manual.",
                "not_found": "Profil santri tidak ditemukan",
                "detector_unavailable": "Detector wajah tidak tersedia di server",
                "invalid_claimed_name": "Nama lengkap tidak ditemukan",
                "duplicate_claimed_name": "Nama lengkap tidak unik, mohon isi lebih spesifik",
            }
            message = error_map.get(info, "Wajah tidak cocok")
            status_code = 400
            retry_map = {
                "no_face": 1200,
                "low_confidence": 1600,
                "ambiguous_face": 1800,
                "not_found": 1800,
                "detector_unavailable": 2500,
                "no_dataset": 3000,
                "not_enough_classes": 3000,
                "invalid_claimed_name": 1800,
                "duplicate_claimed_name": 1800,
            }
            payload = {
                "ok": False,
                "message": message,
                "info": info,
                "retry_after_ms": retry_map.get(info, 1400),
                "can_claim_identity": info in ["ambiguous_face", "invalid_claimed_name", "duplicate_claimed_name"],
            }
            if rejected:
                first = rejected[0]
                location = first.get("location")
                confidence = first.get("confidence")
                if location:
                    payload["location"] = {
                        "top": location[0],
                        "right": location[1],
                        "bottom": location[2],
                        "left": location[3]
                    }
                if confidence is not None:
                    payload["confidence"] = round(float(confidence) * 100, 2)
            payload["detections"] = [
                {
                    "location": {
                        "top": d["location"][0],
                        "right": d["location"][1],
                        "bottom": d["location"][2],
                        "left": d["location"][3],
                    },
                    "confidence": round(float(d["confidence"]) * 100, 2) if d.get("confidence") is not None else None,
                    "info": d.get("info"),
                }
                for d in rejected[:5]
            ]
            return Response(payload, status=status_code)

        status_presensi = "Hadir"
        if telat_start:
            try:
                telat_dt = datetime.datetime.fromisoformat(telat_start)
                if timezone.is_naive(telat_dt):
                    telat_dt = timezone.make_aware(telat_dt)
                diff = (timezone.now() - telat_dt).total_seconds() / 60
                if diff <= 5:
                    status_presensi = "T1"
                elif diff <= 15:
                    status_presensi = "T2"
                else:
                    status_presensi = "T3"
            except Exception as e:
                pass

        attendees = []
        for item in recognized:
            santri = item["santri"]
            location = item["location"]
            confidence = item["confidence"]

            _replace_presensi_record(
                santri=santri,
                tanggal=tanggal,
                sesi=sesi,
                kelas=kelas,
                status_presensi=status_presensi,
                created_by=request.user,
            )

            if kelas and kelas != "Kamu kelas apa?":
                santri.assign_to_kelas(kelas)

            attendees.append({
                "santri": {"id": santri.id, "nama": santri.nama},
                "status": status_presensi,
                "confidence": round(float(confidence) * 100, 2) if confidence is not None else None,
                "location": {
                    "top": location[0],
                    "right": location[1],
                    "bottom": location[2],
                    "left": location[3]
                }
            })

        primary = attendees[0]
        return Response({
            "ok": True,
            "kelas": kelas,
            "count": len(attendees),
            "attendees": attendees,
            # Backward-compatible fields
            "santri": primary["santri"],
            "status": primary["status"],
            "confidence": primary["confidence"],
            "location": primary["location"],
        })
    except Exception as e:
        return Response({"ok": False, "message": f"Error processing: {str(e)}"}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_rekap(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    kelas = request.GET.get('kelas')
    if not start or not end:
        return Response({'ok': False, 'message': 'start & end required'}, status=400)
    
    data = get_rekap_data(start, end, kelas)
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_export_xlsx(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    kelas = request.GET.get('kelas')
    if not start or not end:
        return Response({'ok': False, 'message': 'start & end required'}, status=400)

    data = get_rekap_data(start, end, kelas)
    headers = data['headers']
    putra = data['putra']
    putri = data['putri']

    cols = ['Nama'] + [h['col_key'] for h in headers]
    
    if len(putra) > 0:
        df_putra = pd.DataFrame(putra)[cols]
    else:
        df_putra = pd.DataFrame(columns=cols)
    
    if len(putri) > 0:
        df_putri = pd.DataFrame(putri)[cols]
    else:
        df_putri = pd.DataFrame(columns=cols)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for name, df in [("Putra", df_putra), ("Putri", df_putri)]:
            df.to_excel(writer, index=False, sheet_name=name)

            ws = writer.sheets[name]
            alignment = Alignment(horizontal="center", vertical="center")
            font_bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = font_bold
                cell.alignment = alignment

            fill_colors = {
                "Hadir": "C6EFCE",
                "T1": "FFF2CC", 
                "T2": "FFE699",
                "T3": "FFD966",
                "Izin": "C9DAF8",  
                "-": "F4CCCC"     
            }

            for row in ws.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    val = str(cell.value)
                    if val in fill_colors:
                        cell.fill = PatternFill(start_color=fill_colors[val], end_color=fill_colors[val], fill_type="solid")
                    cell.alignment = alignment

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rekap_presensi.xlsx'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_verify_santri_name(request):
    if not request.user.is_staff:
        return Response({'ok': False, 'message': 'Only pengurus can verify santri'}, status=403)
    
    santri_name = request.data.get('santri_name', '').strip()
    if not santri_name:
        return Response({'ok': False, 'message': 'Nama santri harus diisi'}, status=400)
    
    excel_path = os.path.join(settings.MEDIA_ROOT, 'data_santri', 'data_santri.xlsx')
    if not os.path.exists(excel_path):
        return Response({'ok': False, 'message': 'File data santri tidak ditemukan'}, status=500)
    
    try:
        df = pd.read_excel(excel_path)
        df.columns = df.columns.str.strip().str.lower()
        
        name_column = None
        for col in ['nama', 'name', 'nama santri', 'nama lengkap']:
            if col in df.columns:
                name_column = col
                break
        
        if name_column is None:
            return Response({
                'ok': False, 
                'message': 'Kolom nama tidak ditemukan di file Excel',
                'available_columns': list(df.columns)
            }, status=500)
        
        df[name_column] = df[name_column].astype(str).str.strip()
        found = df[df[name_column].str.lower() == santri_name.lower()]
        
        if found.empty:
            return Response({
                'ok': False, 
                'message': f'Nama "{santri_name}" tidak ditemukan dalam daftar santri resmi',
                'verified': False
            })
        
        reg_code = RegistrationCode.objects.create(
            santri_name=santri_name,
            generated_by=request.user
        )
        
        return Response({
            'ok': True,
            'verified': True,
            'message': f'Santri "{santri_name}" terverifikasi',
            'registration_code': reg_code.code,
            'expires_at': reg_code.expires_at.isoformat()
        })
        
    except Exception as e:
        return Response({
            'ok': False,
            'message': f'Error reading Excel file: {str(e)}'
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_list_registration_codes(request):
    if not request.user.is_staff:
        return Response({'ok': False, 'message': 'Only pengurus can view codes'}, status=403)
    
    codes = RegistrationCode.objects.all().order_by('-created_at')
    data = [{
        'id': code.id,
        'code': code.code,
        'santri_name': code.santri_name,
        'used': code.used,
        'is_valid': code.is_valid(),
        'created_at': code.created_at,
        'expires_at': code.expires_at,
        'generated_by': code.generated_by.username if code.generated_by else None,
        'used_by': code.used_by.username if code.used_by else None
    } for code in codes]
    
    return Response({'ok': True, 'data': data})
